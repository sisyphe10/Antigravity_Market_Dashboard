# -*- coding: utf-8 -*-
"""미국·한국 회사채 발행액 월간 시계열 수집 → dataset.csv

시리즈 2종 (합계만, 2026-08-07 사용자 확정):
  미 회사채 발행액   ($B, 소수 1자리)  dtype=BOND_ISSUANCE_US
  회사채 발행액      (억원, 정수)      dtype=BOND_ISSUANCE_KR
dtype 을 소스별로 분리 = check_data_freshness 개별 추적 (SiliconData 3종과 동일 원칙 —
공유 타입이면 한 원천이 죽어도 가려진다).

## 미국 원천 — SIFMA US Fixed Income Securities Statistics xlsx (직링크)
- Issuance 시트의 Corporates 열 (all non-convertible/convertible debt, MTNs, Yankee bonds).
- ★파일의 월별 데이터는 최근 13개월 롤링 창뿐 → 매 run 창 전체를 재파싱해 개정치 upsert.
- ★과거 이력은 Wayback 스냅숏 6개(2021~2025)를 --backfill 로 1회 스티칭:
  2019-01~2026-07 복원, 단 2022-10~2023-04 7개월은 아카이브 공백 = 영구 결측.
  (회사채 전용 파일(1996~)은 2026년 현재 HubSpot 폼 뒤 게이트 — 직링크·아카이브 모두 없음)
- 스냅숏 구간 값은 당시 공표 빈티지(이후 소급 개정 미반영) — 2025 연간 합이 최신
  공표치와 ~0.7% 어긋나는 이유. 현행 13개월 창 구간은 매 run 최신 개정으로 수렴.
- 구층 레이아웃 대응: 헤더 'Corporates'(현행) | 'Corporate Debt'(2021~22), 시트명 'Issuance'.

## 한국 원천 — 금투협 채권정보센터 발행통계 기간별 (2026-08-07 의미 계약 확정)
- POST https://www.kofiabond.or.kr/proframeWeb/XMLSERVICES/ (proframe XML, 세션 불필요)
  BIS-KOFIABOND / BISIssStatisSrchSO / listTrm, val1·val2 = YYYYMMDD 기간.
- 응답: 채권종류별 행(국채/지방채/특수채/통안증권/은행채/기타금융채/회사채/ABS).
  ★'회사채' = 일반회사채 (은행채·기타금융채·ABS 별도 행 — 금감원 직접금융 총계(금융채·ABS
  포함)와 모집단이 다르다. 임의 합산 금지, 금감원 수치를 앵커로 쓰지 말 것).
- 행 판정 = val1=='회사채' AND 영문명 'Corporate' 동시 일치. 값 = val2(발행액), 단위 억원
  (메타 displayUnitNm='억원,%' 확정). 이력 2006-01~ (2005 이전은 행 없음).
- 월별 = 월초~월말 1콜. 진행 중인 달은 부분값이라 제외(월말 < 오늘 인 달만),
  최근 4개 완료월은 매 run 재조회해 D+1 지연·정정을 self-heal.

사용:
  venv/bin/python3 execution/fetch_corp_bond_issuance.py            # 증분 (매일 잡)
  venv/bin/python3 execution/fetch_corp_bond_issuance.py --backfill # 최초 1회 전체 백필
  venv/bin/python3 execution/fetch_corp_bond_issuance.py --dry-run
"""
import calendar
import csv
import io
import os
import re
import sys
import time
from datetime import date, timedelta

import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

CSV_PATH = 'dataset.csv'
TIMEOUT = 60
SLEEP = 0.4

US_NAME = '미 회사채 발행액'
US_DTYPE = 'BOND_ISSUANCE_US'
KR_NAME = '회사채 발행액'
KR_DTYPE = 'BOND_ISSUANCE_KR'

SIFMA_URL = ('https://www.sifma.org/wp-content/uploads/2021/02/'
             'US-Fixed-Income-Securities-Statistics-SIFMA.xlsx')
# Wayback 스냅숏 (digest-unique 전수, 2026-08-07 확인) — 백필 전용
SIFMA_WAYBACK_TS = ['20211102062659', '20220701225656', '20220904054230',
                    '20221024141204', '20240613004826', '20250723005053']
US_LO, US_HI = 0.0, 1000.0          # $B/월
US_MAX_LAG_DAYS = 75                # 월별 공표 랙 상한 (통상 다음달 초·간혹 지연)
# 앵커: 파싱 창에 해당 월이 있을 때만 검사 (창이 지나가면 자동 소멸) — 소급 개정 감안 10%
US_ANCHORS = {'2025-12-31': 69.8, '2026-06-30': 276.1}
US_ANCHOR_RTOL = 0.10

KOFIA_URL = 'https://www.kofiabond.or.kr/proframeWeb/XMLSERVICES/'
KOFIA_BODY = ("<?xml version='1.0' encoding='utf-8'?><message><proframeHeader>"
              "<pfmAppName>BIS-KOFIABOND</pfmAppName>"
              "<pfmSvcName>BISIssStatisSrchSO</pfmSvcName>"
              "<pfmFnName>listTrm</pfmFnName></proframeHeader><systemHeader>"
              "</systemHeader><BISComDspDatDTO><val1>{}</val1><val2>{}</val2>"
              "</BISComDspDatDTO></message>")
KR_START = (2006, 1)                # 회사채 행 존재 시작 (2005 이전 없음, 실측)
KR_LO, KR_HI = 0.0, 1_000_000.0     # 억원/월
KR_HEAL_MONTHS = 4                  # 매 run 재조회해 self-heal 하는 최근 완료월 수
KR_ANCHORS = {'2010-01-31': 28796.0, '2020-07-31': 87868.0}   # 조회 범위에 있을 때만
KR_ANCHOR_RTOL = 0.03

RAW_DIR = os.path.join(os.path.expanduser('~'), 'datalake', 'raw', 'sifma')


class SrcError(Exception):
    pass


def month_end(y, m):
    return date(y, m, calendar.monthrange(y, m)[1])


def add_months(y, m, k):
    m = m + k
    y += (m - 1) // 12
    return y, (m - 1) % 12 + 1


def fmt(v, nd):
    out = f'{v:.{nd}f}'
    if '.' in out:
        out = out.rstrip('0').rstrip('.')
    return '0' if out in ('-0', '') else out


def http_get(url):
    import urllib.request
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
        return r.read()


def http_post(url, body):
    import urllib.request
    req = urllib.request.Request(url, data=body.encode('utf-8'),
                                 headers={'Content-Type': 'application/xml',
                                          'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
        return r.read().decode('utf-8')


# ─────────────────────────── 미국: SIFMA ───────────────────────────

def sifma_extract_monthly(blob):
    """xlsx bytes → {관측월 말일 date: Corporates $B}. 위치 하드코딩 금지 —
    Issuance 시트·Corporates 헤더를 문자열로 탐색, datetime 라벨 행만 월별로 인정."""
    if not blob.startswith(b'PK\x03\x04') or len(blob) < 50000:
        raise SrcError('xlsx 아님 (폼 게이트/HTML 응답 의심)')
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        sheet = None
        for name in wb.sheetnames:
            if name.strip().lower() == 'issuance':
                sheet = wb[name]
                break
        if sheet is None:
            raise SrcError(f'Issuance 시트 없음: {wb.sheetnames}')
        corp_col, pts = None, {}
        for row in sheet.iter_rows(values_only=True):
            cells = list(row)
            if corp_col is None:
                for i, c in enumerate(cells):
                    if isinstance(c, str) and c.strip() in ('Corporates', 'Corporate Debt'):
                        labels = [str(x).strip() for x in cells if isinstance(x, str)]
                        if 'UST' in labels or 'Treasury' in labels:
                            corp_col = i
                        break
                continue
            label = cells[0] if cells else None
            if hasattr(label, 'year') and hasattr(label, 'month'):   # datetime = 월별 행
                val = cells[corp_col] if corp_col < len(cells) else None
                if isinstance(val, (int, float)):
                    v = float(val)
                    if not (US_LO < v < US_HI):
                        raise SrcError(f'값 범위 이탈 {label:%Y-%m} → {v}')
                    pts[month_end(label.year, label.month)] = round(v, 1)
        if corp_col is None:
            raise SrcError('Corporates 헤더 없음 (레이아웃 변경 의심)')
        return pts
    finally:
        wb.close()


def save_raw(blob, tag):
    """맥미니(@lake 존재 시)에서만 원본 영구보존 — GHA 러너에선 조용히 skip."""
    try:
        if os.path.isdir(os.path.dirname(RAW_DIR)):
            os.makedirs(RAW_DIR, exist_ok=True)
            path = os.path.join(RAW_DIR, f'sifma_fi_{tag}.xlsx')
            if not os.path.exists(path):
                with open(path, 'wb') as f:
                    f.write(blob)
    except OSError as e:
        print(f'    · 원본 보존 실패(무해): {e}')


def fetch_us(backfill, today):
    merged = {}
    if backfill:
        for ts in SIFMA_WAYBACK_TS:
            blob = http_get(f'http://web.archive.org/web/{ts}id_/{SIFMA_URL}')
            pts = sifma_extract_monthly(blob)
            save_raw(blob, f'wayback_{ts}')
            print(f'    · wayback {ts}: {len(pts)}개월 ({min(pts)}~{max(pts)})')
            merged.update(pts)          # 나중(최신) 스냅숏이 개정치로 덮어씀
            time.sleep(1.5)
    blob = http_get(SIFMA_URL)
    pts = sifma_extract_monthly(blob)
    if len(pts) < 12:
        raise SrcError(f'현행 파일 월별 {len(pts)}개월 — 창 축소/레이아웃 변경 의심')
    if (today - max(pts)).days > US_MAX_LAG_DAYS:
        raise SrcError(f'최신월 {max(pts)} — 랙 {(today - max(pts)).days}일 초과')
    for iso, want in US_ANCHORS.items():
        d = date.fromisoformat(iso)
        got = pts.get(d)
        if got is not None and abs(got - want) > want * US_ANCHOR_RTOL:
            raise SrcError(f'앵커 불일치 {iso}: 기대 {want}±10%, 실제 {got}')
    save_raw(blob, f'{max(pts):%Y%m}')
    merged.update(pts)
    return merged


# ─────────────────────────── 한국: 금투협 ───────────────────────────

def kofia_month(y, m):
    """해당 월 회사채 발행액(억원). 회사채 행 없으면 None (2006 이전 정상)."""
    d1 = f'{y}{m:02d}01'
    d2 = f'{y}{m:02d}{calendar.monthrange(y, m)[1]}'
    raw = http_post(KOFIA_URL, KOFIA_BODY.format(d1, d2))
    rows = re.findall(r'<BISComDspDatDTO>(.*?)</BISComDspDatDTO>', raw, re.S)
    if not rows:
        raise SrcError(f'{y}-{m:02d}: 응답에 데이터 행 없음 (계약 변경 의심)')
    for row in rows:
        vals = [v[1].strip() for v in re.findall(r'<(val\d+)>(.*?)</\1>', row, re.S)]
        if not vals or vals[0] != '회사채':
            continue
        if not any(v == 'Corporate' for v in vals):
            raise SrcError(f'{y}-{m:02d}: 회사채 행 영문명 불일치 (분류 변경 의심)')
        try:
            v = float(vals[1].replace(',', ''))
        except (IndexError, ValueError):
            raise SrcError(f'{y}-{m:02d}: 발행액 파싱 실패')
        if not (KR_LO <= v < KR_HI):
            raise SrcError(f'{y}-{m:02d}: 값 범위 이탈 → {v}')
        return v
    return None


def fetch_kr(backfill, today):
    # 완료월만: 월말 < 오늘 (진행 중인 달은 부분값)
    last = add_months(today.year, today.month, -1)
    if month_end(*last) >= today:
        last = add_months(*last, -1)
    if backfill:
        start = KR_START
    else:
        start = add_months(*last, -(KR_HEAL_MONTHS - 1))
    pts, missing = {}, []
    y, m = start
    while (y, m) <= last:
        v = kofia_month(y, m)
        if v is None:
            missing.append(f'{y}-{m:02d}')
        else:
            pts[month_end(y, m)] = v
        y, m = add_months(y, m, 1)
        time.sleep(SLEEP)
    if not pts:
        raise SrcError('유효 관측 0건')
    if not backfill and missing:
        raise SrcError(f'최근 완료월 결측: {", ".join(missing)}')
    if backfill and missing:
        print(f'    · 회사채 행 없는 달 {len(missing)}건 skip (초기 구간): '
              f'{", ".join(missing[:6])}{" ..." if len(missing) > 6 else ""}')
    for iso, want in KR_ANCHORS.items():
        d = date.fromisoformat(iso)
        got = pts.get(d)
        if got is not None and abs(got - want) > want * KR_ANCHOR_RTOL:
            raise SrcError(f'앵커 불일치 {iso}: 기대 {want}±3%, 실제 {got}')
    return pts


# ─────────────────────────── 적재 ───────────────────────────

def main():
    dry = '--dry-run' in sys.argv
    backfill = '--backfill' in sys.argv
    today = date.today()

    header = ['날짜', '제품명', '가격', '데이터 타입']
    all_rows = []
    if os.path.exists(CSV_PATH):
        with open(CSV_PATH, encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            h = next(reader, None)
            if h:
                header = h
            all_rows = [row for row in reader if row]
    index = {}
    for i, row in enumerate(all_rows):
        if len(row) >= 2:
            index[(row[0], row[1])] = i

    sources = [
        ('미국 SIFMA', US_NAME, US_DTYPE, 1, fetch_us),
        ('한국 금투협', KR_NAME, KR_DTYPE, 0, fetch_kr),
    ]
    new_rows, healed, failed = [], 0, []
    for src_label, name, dtype, nd, fn in sources:
        try:
            pts = fn(backfill, today)
        except Exception as e:      # 소스 단위 격리 — 성공한 쪽은 저장
            failed.append(src_label)
            print(f'  ⚠️ {src_label}: {type(e).__name__ if not isinstance(e, SrcError) else ""}{e}')
            continue
        added = 0
        for stamp in sorted(pts):
            if stamp >= today:      # 미래·당일 스탬프 방어
                continue
            s = fmt(pts[stamp], nd)
            k = (stamp.isoformat(), name)
            if k in index:
                row = all_rows[index[k]]
                if row[2] != s:
                    try:
                        same = abs(float(row[2].replace(',', '')) - float(s)) < 1e-9
                    except ValueError:
                        same = False
                    if not same:
                        row[2] = s
                        healed += 1
                while len(row) < 4:
                    row.append('')
                if row[3] != dtype:
                    row[3] = dtype
                    healed += 1
            else:
                row = [k[0], name, s, dtype]
                all_rows.append(row)
                index[k] = len(all_rows) - 1
                new_rows.append(row)
                added += 1
        print(f'  ✓ {src_label} {name}: {len(pts)}건 ({min(pts)}~{max(pts)}), '
              f'최신 {fmt(pts[max(pts)], nd)}, 신규 {added}')

    print(f'\n회사채 발행액 수집: {len(sources) - len(failed)}/{len(sources)} 성공, '
          f'신규 {len(new_rows)}건, 개정 {healed}건'
          + (' [dry-run, 미기록]' if dry else ''))
    if failed:
        print(f'실패 소스: {", ".join(failed)}')

    if not dry:
        if healed:
            with open(CSV_PATH, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f)
                w.writerow(header)
                w.writerows(all_rows)
        elif new_rows:
            write_header = not os.path.exists(CSV_PATH)
            with open(CSV_PATH, 'a', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f)
                if write_header:
                    w.writerow(header)
                w.writerows(new_rows)

    return 1 if failed else 0     # ★부분 실패도 실패 — 감시 신호


if __name__ == '__main__':
    sys.exit(main())
