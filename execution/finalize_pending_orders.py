"""orders/pending_orders.json → Wrap_NAV.xlsx NEW 시트 반영

매일 16:00 KST GHA에서 호출 (workflow_dispatch 수동 가능):
1. orders/pending_orders.json 로드
2. 모든 과거 날짜 entry 처리 (오늘 포함, 미래 제외)
   - 카드별 newSheetTargets (broker, product) × stocks 곱 만큼 행 생성
   - 같은 (날짜, broker, product) 기존 행은 먼저 제거 (덮어쓰기)
3. 처리 완료된 날짜 entry는 pending_orders.json에서 삭제

Wrap_NAV.xlsx, orders/pending_orders.json 둘 다 commit 됨.
"""
import json
import os
import sys
from datetime import datetime, timezone, timedelta
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WRAP_NAV = os.path.join(ROOT, 'Wrap_NAV.xlsx')
PENDING = os.path.join(ROOT, 'orders', 'pending_orders.json')

KST = timezone(timedelta(hours=9))


def main():
    if not os.path.exists(PENDING):
        print('⚪ pending_orders.json 없음 — 처리할 ORDER 없음')
        return 0
    with open(PENDING, encoding='utf-8') as f:
        pending = json.load(f)
    if not isinstance(pending, dict) or not pending:
        print('⚪ pending_orders.json 비어있음')
        return 0

    if not os.path.exists(WRAP_NAV):
        print('❌ Wrap_NAV.xlsx 없음')
        return 1
    wb = openpyxl.load_workbook(WRAP_NAV)
    if 'NEW' not in wb.sheetnames:
        print('❌ NEW 시트 없음')
        return 1
    ws = wb['NEW']

    today_kst = datetime.now(tz=KST).strftime('%Y-%m-%d')
    sorted_dates = sorted(pending.keys())

    processed_dates = []
    rows_added_total = 0
    for date_str in sorted_dates:
        if date_str > today_kst:
            print(f'⏭ {date_str} 미래 날짜 — skip (보류)')
            continue
        cards = pending[date_str]
        if not isinstance(cards, dict) or not cards:
            processed_dates.append(date_str)
            continue
        print(f'=== {date_str}: 카드 {len(cards)}개 ===')

        date_dt = datetime.strptime(date_str, '%Y-%m-%d')

        for card_name, card_entry in cards.items():
            targets = card_entry.get('targets', [])
            stocks = card_entry.get('stocks', [])
            if not targets or not stocks:
                print(f'  ⚠️ {card_name}: targets/stocks 비어있음, skip')
                continue

            # 같은 (날짜, broker, product) 기존 행 제거
            rows_to_delete = []
            for r in range(2, ws.max_row + 1):
                date_v = ws.cell(row=r, column=1).value
                broker_v = ws.cell(row=r, column=2).value
                product_v = ws.cell(row=r, column=3).value
                if not broker_v or not product_v:
                    continue
                row_date_str = date_v.strftime('%Y-%m-%d') if isinstance(date_v, datetime) else str(date_v)[:10]
                if row_date_str != date_str:
                    continue
                for t in targets:
                    if broker_v == t['broker'] and product_v == t['product']:
                        rows_to_delete.append(r)
                        break
            for r in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(r)
            if rows_to_delete:
                print(f'  {card_name}: 기존 {len(rows_to_delete)}행 제거')

            added = 0
            for t in targets:
                for s in stocks:
                    ws.append([date_dt, t['broker'], t['product'], s.get('sector', ''),
                               str(s.get('code', '')), s.get('name', ''), s.get('weight', 0)])
                    added += 1
            rows_added_total += added
            print(f'  ✓ {card_name}: {len(stocks)}종목 × {len(targets)}상품 = {added}행 추가')

        processed_dates.append(date_str)

    if not processed_dates:
        print('처리할 날짜 없음')
        return 0

    wb.save(WRAP_NAV)
    print(f'\n✅ Wrap_NAV.xlsx 저장 (총 {rows_added_total}행 추가)')

    for d in processed_dates:
        if d in pending:
            del pending[d]
    with open(PENDING, 'w', encoding='utf-8') as f:
        json.dump(pending, f, ensure_ascii=False, indent=2)
    print(f'✅ pending_orders.json 에서 {len(processed_dates)}개 날짜 entry 제거')
    return 0


if __name__ == '__main__':
    sys.exit(main())
