"""orders/aum_pending.json → Wrap_NAV.xlsx AUM 시트 반영

매일 16:00 KST GHA에서 호출 (finalize_orders와 같은 workflow):
1. orders/aum_pending.json 로드
2. 모든 과거 날짜 entry 처리 (오늘 포함, 미래 제외)
   - 같은 (날짜, broker, product) 기존 행 제거 후 새 행 추가 (덮어쓰기)
3. 처리 완료된 날짜 entry는 aum_pending.json에서 삭제

Wrap_NAV.xlsx, orders/aum_pending.json 둘 다 commit 됨.
"""
import json
import os
import sys
from datetime import datetime, timezone, timedelta
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WRAP_NAV = os.path.join(ROOT, 'Wrap_NAV.xlsx')
PENDING = os.path.join(ROOT, 'orders', 'aum_pending.json')

KST = timezone(timedelta(hours=9))


def main():
    if not os.path.exists(PENDING):
        print('⚪ aum_pending.json 없음 — 처리할 AUM 없음')
        return 0
    with open(PENDING, encoding='utf-8') as f:
        pending = json.load(f)
    if not isinstance(pending, dict) or not pending:
        print('⚪ aum_pending.json 비어있음')
        return 0

    if not os.path.exists(WRAP_NAV):
        print('❌ Wrap_NAV.xlsx 없음')
        return 1
    wb = openpyxl.load_workbook(WRAP_NAV)
    if 'AUM' not in wb.sheetnames:
        print('❌ AUM 시트 없음')
        return 1
    ws = wb['AUM']

    today_kst = datetime.now(tz=KST).strftime('%Y-%m-%d')
    sorted_dates = sorted(pending.keys())

    processed_dates = []
    rows_added_total = 0
    for date_str in sorted_dates:
        if date_str > today_kst:
            print(f'⏭ {date_str} 미래 날짜 — skip')
            continue
        entries = pending[date_str]
        if not isinstance(entries, list) or not entries:
            processed_dates.append(date_str)
            continue
        print(f'=== {date_str}: {len(entries)} entries ===')

        date_dt = datetime.strptime(date_str, '%Y-%m-%d')

        # 같은 (날짜, broker, product) 기존 행 제거
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            date_v = ws.cell(row=r, column=1).value
            broker_v = ws.cell(row=r, column=2).value
            product_v = ws.cell(row=r, column=3).value
            if not broker_v or not product_v:
                continue
            row_date_str = date_v.strftime('%Y-%m-%d') if isinstance(date_v, datetime) else str(date_v)[:10]
            if row_date_str != date_str:
                continue
            for e in entries:
                if broker_v == e['broker'] and product_v == e['product']:
                    rows_to_delete.append(r)
                    break
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)
        if rows_to_delete:
            print(f'  기존 {len(rows_to_delete)}행 제거')

        for e in entries:
            ws.append([date_dt, e['broker'], e['product'], int(e['aum'])])
            rows_added_total += 1
        print(f'  ✓ {len(entries)}행 추가')
        processed_dates.append(date_str)

    if not processed_dates:
        print('처리할 날짜 없음')
        return 0

    wb.save(WRAP_NAV)
    print(f'\n✅ Wrap_NAV.xlsx 저장 (총 {rows_added_total}행 추가)')

    for d in processed_dates:
        if d in pending:
            del pending[d]
    with open(PENDING, 'w', encoding='utf-8') as f:
        json.dump(pending, f, ensure_ascii=False, indent=2)
    print(f'✅ aum_pending.json 에서 {len(processed_dates)}개 날짜 entry 제거')
    return 0


if __name__ == '__main__':
    sys.exit(main())
