"""전체 620종목 WICS 소분류 통일 매핑 → WICS_분류.xlsx 최종본 시트"""
import requests, json, sys, openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

# === 1. 데이터 수집 ===
url = 'https://sheets.googleapis.com/v4/spreadsheets/1KR9RJN53G-yJtnowQbg5bcAiIBfrkIeNqN_PO2UOCTM/values/universe?key=AIzaSyCHPiRby5FVAIKDwneZHy1KGl3SfycjZEw'
r = requests.get(url, timeout=30)
rows = r.json().get('values', [])[1:]
print(f'전체 {len(rows)}종목')

# WICS 매핑 (KRW)
with open('wics_mapping.json', encoding='utf-8') as f:
    wics = json.load(f)['mapping']

# yfinance → WICS 매핑
YF_TO_WICS = {
    'Oil & Gas Midstream': '에너지', 'Uranium': '에너지', 'Solar': '에너지',
    'Oil & Gas E&P': '에너지', 'Oil & Gas Integrated': '에너지',
    'Chemicals': '소재', 'Other Industrial Metals & Mining': '소재',
    'Aerospace & Defense': '자본재', 'Conglomerates': '자본재',
    'Engineering & Construction': '자본재', 'Specialty Industrial Machinery': '자본재',
    'Electrical Equipment & Parts': '자본재', 'Building Products & Equipment': '자본재',
    'Airlines': '운송',
    'Auto Manufacturers': '자동차와부품', 'Auto Parts': '자동차와부품',
    'Apparel Manufacturing': '내구소비재와의류', 'Apparel Retail': '내구소비재와의류',
    'Footwear & Accessories': '내구소비재와의류', 'Luxury Goods': '내구소비재와의류',
    'Leisure': '내구소비재와의류', 'Furnishings, Fixtures & Appliances': '내구소비재와의류',
    'Restaurants': '호텔,레스토랑,레저등',
    'Discount Stores': '소매(유통)', 'Home Improvement Retail': '소매(유통)',
    'Specialty Retail': '소매(유통)', 'Internet Retail': '소매(유통)',
    'Beverages - Non-Alcoholic': '식품,음료,담배', 'Tobacco': '식품,음료,담배',
    'Household & Personal Products': '가정용품과개인용품',
    'Medical Devices': '건강관리장비와서비스', 'Medical Instruments & Supplies': '건강관리장비와서비스',
    'Diagnostics & Research': '건강관리장비와서비스',
    'Drug Manufacturers - General': '제약과생물공학',
    'Drug Manufacturers - Specialty & Generic': '제약과생물공학',
    'Biotechnology': '제약과생물공학',
    'Asset Management': '다각화된금융', 'Credit Services': '다각화된금융',
    'Financial Data & Stock Exchanges': '다각화된금융', 'Capital Markets': '다각화된금융',
    'Software - Application': '소프트웨어와서비스',
    'Software - Infrastructure': '소프트웨어와서비스',
    'Information Technology Services': '소프트웨어와서비스',
    'Computer Hardware': '기술하드웨어와장비', 'Communication Equipment': '기술하드웨어와장비',
    'Electronic Components': '기술하드웨어와장비',
    'Semiconductors': '반도체와반도체장비',
    'Semiconductor Equipment & Materials': '반도체와반도체장비',
    'Consumer Electronics': '전자와 전기제품',
    'Telecom Services': '전기통신서비스',
    'Internet Content & Information': '미디어와엔터테인먼트',
    'Entertainment': '미디어와엔터테인먼트',
    'Electronic Gaming & Multimedia': '미디어와엔터테인먼트',
    'Advertising Agencies': '미디어와엔터테인먼트',
    'Utilities - Independent Power Producers': '유틸리티',
    'Utilities - Regulated Electric': '유틸리티',
}

# 해외 티커 → yfinance 심볼
YF_TICKER = {
    'TPE:2408': '2408.TW', 'TPE:2454': '2454.TW', 'TPE:6515': '6515.TW',
    'HKG:0700': '0700.HK', 'HKG:1810': '1810.HK', 'HKG:9992': '9992.HK',
    'HKG:1913': '1913.HK', 'HKG:81211': '1211.HK',
    'AMS:BESI': 'BESI.AS', 'EPA:ETL': 'ETL.PA',
    'TSE:ATZ': 'ATZ.TO', 'ETR:ADS': 'ADS.DE', 'ETR:RHM': 'RHM.DE',
}

# === 2. 전체 매핑 ===
import yfinance as yf

results = []
foreign_done = 0

for row in rows:
    currency = row[1] if len(row) > 1 else ''
    old_sector = row[2] if len(row) > 2 else ''
    ticker = row[3] if len(row) > 3 else ''
    name = row[4] if len(row) > 4 else ''

    if currency == 'KRW':
        code = ticker.split(':')[-1] if ':' in ticker else ticker
        new_sector = wics.get(code, old_sector)
        results.append((currency, ticker, name, old_sector, new_sector, ''))
    else:
        yf_sym = YF_TICKER.get(ticker, '')
        if not yf_sym:
            yf_sym = ticker.split(':')[-1] if ':' in ticker else ticker
        try:
            info = yf.Ticker(yf_sym).info
            industry = info.get('industry', '')
            new_sector = YF_TO_WICS.get(industry, industry) if industry else old_sector
        except:
            industry = ''
            new_sector = old_sector
        results.append((currency, ticker, name, old_sector, new_sector, industry))
        foreign_done += 1
        if foreign_done % 30 == 0:
            print(f'  해외 {foreign_done}종목 처리...')

changed = sum(1 for r in results if r[3] != r[4])
print(f'전체 {len(results)}종목, {changed}종목 변경')

# === 3. 엑셀 ===
wb = openpyxl.load_workbook('WICS_분류.xlsx')

if '최종본' in wb.sheetnames:
    del wb['최종본']

ws = wb.create_sheet('최종본')

headers = ['통화', '티커', '기업명', '현재 섹터', 'WICS 소분류', 'industry(원문)']
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
changed_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin
    cell.alignment = Alignment(horizontal='center')

for i, (cur, ticker, name, old, new, en) in enumerate(results, 2):
    ws.cell(row=i, column=1, value=cur).border = thin
    ws.cell(row=i, column=2, value=ticker).border = thin
    ws.cell(row=i, column=3, value=name).border = thin
    ws.cell(row=i, column=4, value=old).border = thin
    c = ws.cell(row=i, column=5, value=new)
    c.border = thin
    if old != new:
        c.fill = changed_fill
        c.font = Font(bold=True)
    ws.cell(row=i, column=6, value=en).border = thin

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 38
ws.auto_filter.ref = f'A1:F{len(results)+1}'

# 분류 매핑 시트
if '분류 매핑' in wb.sheetnames:
    del wb['분류 매핑']

ws_map = wb.create_sheet('분류 매핑', 0)
mapping_data = [
    ('에너지', 'Oil & Gas Midstream, Uranium, Solar'),
    ('소재', 'Chemicals, Other Industrial Metals & Mining'),
    ('자본재', 'Aerospace & Defense, Conglomerates, Engineering & Construction, Specialty Industrial Machinery, Electrical Equipment & Parts, Building Products & Equipment'),
    ('상업서비스와공급품', '—'),
    ('운송', 'Airlines'),
    ('자동차와부품', 'Auto Manufacturers, Auto Parts'),
    ('내구소비재와의류', 'Apparel Manufacturing/Retail, Footwear & Accessories, Luxury Goods, Leisure, Furnishings'),
    ('호텔,레스토랑,레저등', 'Restaurants'),
    ('소매(유통)', 'Discount Stores, Home Improvement Retail, Specialty Retail, Internet Retail'),
    ('교육서비스', '—'),
    ('식품과기본식료품소매', '—'),
    ('식품,음료,담배', 'Beverages - Non-Alcoholic, Tobacco'),
    ('가정용품과개인용품', 'Household & Personal Products'),
    ('건강관리장비와서비스', 'Medical Devices, Medical Instruments & Supplies, Diagnostics & Research'),
    ('제약과생물공학', 'Drug Manufacturers (General/Specialty), Biotechnology'),
    ('은행', '—'),
    ('증권', '—'),
    ('다각화된금융', 'Asset Management, Credit Services, Financial Data & Stock Exchanges, Capital Markets'),
    ('보험', '—'),
    ('부동산', '—'),
    ('소프트웨어와서비스', 'Software - Application/Infrastructure, IT Services'),
    ('기술하드웨어와장비', 'Computer Hardware, Communication Equipment, Electronic Components'),
    ('반도체와반도체장비', 'Semiconductors, Semiconductor Equipment & Materials'),
    ('전자와 전기제품', 'Consumer Electronics'),
    ('디스플레이', '—'),
    ('전기통신서비스', 'Telecom Services'),
    ('미디어와엔터테인먼트', 'Internet Content & Information, Entertainment, Gaming, Advertising'),
    ('유틸리티', 'Utilities (Independent Power Producers, Regulated Electric)'),
]

for col, h in enumerate(['WICS 소분류', 'yfinance industry (매핑)'], 1):
    cell = ws_map.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin
    cell.alignment = Alignment(horizontal='center')

for i, (w, y) in enumerate(mapping_data, 2):
    ws_map.cell(row=i, column=1, value=w).border = thin
    ws_map.cell(row=i, column=1).font = Font(bold=True)
    ws_map.cell(row=i, column=2, value=y).border = thin

ws_map.column_dimensions['A'].width = 24
ws_map.column_dimensions['B'].width = 80

wb.save('WICS_분류.xlsx')
print(f'완료! 시트: {wb.sheetnames}')
