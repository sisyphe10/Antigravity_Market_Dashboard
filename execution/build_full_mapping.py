"""
전체 섹터 매핑 엑셀 생성
- 국내: WICS 소분류 2,550종목
- S&P 500: Wikipedia GICS → WICS 매핑
- NASDAQ 시총 상위: yfinance → WICS 매핑
"""
import requests, json, sys, io, time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import yfinance as yf

sys.stdout.reconfigure(encoding='utf-8')

thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


# === GICS Sub-Industry → WICS 소분류 매핑 ===
GICS_TO_WICS = {
    # 에너지
    'Oil & Gas Exploration & Production': '에너지', 'Oil & Gas Refining & Marketing': '에너지',
    'Integrated Oil & Gas': '에너지', 'Oil & Gas Storage & Transportation': '에너지',
    'Oil & Gas Equipment & Services': '에너지', 'Oil & Gas Drilling': '에너지',
    'Coal & Consumable Fuels': '에너지',
    # 소재
    'Commodity Chemicals': '소재', 'Specialty Chemicals': '소재', 'Diversified Chemicals': '소재',
    'Industrial Gases': '소재', 'Fertilizers & Agricultural Chemicals': '소재',
    'Gold': '소재', 'Copper': '소재', 'Diversified Metals & Mining': '소재', 'Steel': '소재',
    'Aluminum': '소재', 'Silver': '소재', 'Paper & Plastic Packaging Products & Materials': '소재',
    'Metal, Glass & Plastic Containers': '소재', 'Paper Products': '소재',
    'Construction Materials': '소재', 'Forest Products': '소재',
    # 자본재
    'Aerospace & Defense': '자본재', 'Industrial Conglomerates': '자본재',
    'Building Products': '자본재', 'Construction & Engineering': '자본재',
    'Electrical Components & Equipment': '자본재', 'Heavy Electrical Equipment': '자본재',
    'Industrial Machinery & Supplies & Components': '자본재',
    'Construction Machinery & Heavy Transportation Equipment': '자본재',
    'Agricultural & Farm Machinery': '자본재', 'Trading Companies & Distributors': '자본재',
    # 상업서비스와공급품
    'Environmental & Facilities Services': '상업서비스와공급품',
    'Office Services & Supplies': '상업서비스와공급품',
    'Diversified Support Services': '상업서비스와공급품',
    'Security & Alarm Services': '상업서비스와공급품',
    'Human Resource & Employment Services': '상업서비스와공급품',
    'Research & Consulting Services': '상업서비스와공급품',
    'Data Processing & Outsourced Services': '상업서비스와공급품',
    # 운송
    'Air Freight & Logistics': '운송', 'Passenger Airlines': '운송', 'Airlines': '운송',
    'Marine Transportation': '운송', 'Railroads': '운송', 'Cargo Ground Transportation': '운송',
    'Passenger Ground Transportation': '운송',
    # 자동차와부품
    'Automobile Manufacturers': '자동차와부품', 'Automotive Parts & Equipment': '자동차와부품',
    'Tires & Rubber': '자동차와부품', 'Motor Vehicle Parts': '자동차와부품',
    # 내구소비재와의류
    'Apparel, Accessories & Luxury Goods': '내구소비재와의류', 'Footwear': '내구소비재와의류',
    'Textiles': '내구소비재와의류', 'Homebuilding': '내구소비재와의류',
    'Home Furnishings': '내구소비재와의류', 'Household Appliances': '내구소비재와의류',
    'Housewares & Specialties': '내구소비재와의류', 'Leisure Products': '내구소비재와의류',
    # 호텔,레스토랑,레저등
    'Hotels, Resorts & Cruise Lines': '호텔,레스토랑,레저등', 'Restaurants': '호텔,레스토랑,레저등',
    'Leisure Facilities': '호텔,레스토랑,레저등', 'Casinos & Gaming': '호텔,레스토랑,레저등',
    # 소매(유통)
    'Internet & Direct Marketing Retail': '소매(유통)', 'Broadline Retail': '소매(유통)',
    'Apparel Retail': '소매(유통)', 'Computer & Electronics Retail': '소매(유통)',
    'Home Improvement Retail': '소매(유통)', 'Automotive Retail': '소매(유통)',
    'Homefurnishing Retail': '소매(유통)', 'Specialty Stores': '소매(유통)',
    'General Merchandise Stores': '소매(유통)', 'Distributors': '소매(유통)',
    # 식품과기본식료품소매
    'Consumer Staples Merchandise Retail': '식품과기본식료품소매',
    'Food Distributors': '식품과기본식료품소매', 'Food Retail': '식품과기본식료품소매',
    'Drug Retail': '식품과기본식료품소매',
    # 식품,음료,담배
    'Packaged Foods & Meats': '식품,음료,담배', 'Soft Drinks & Non-alcoholic Beverages': '식품,음료,담배',
    'Brewers': '식품,음료,담배', 'Distillers & Vintners': '식품,음료,담배',
    'Tobacco': '식품,음료,담배', 'Agricultural Products & Services': '식품,음료,담배',
    # 가정용품과개인용품
    'Household Products': '가정용품과개인용품', 'Personal Care Products': '가정용품과개인용품',
    # 건강관리장비와서비스
    'Health Care Equipment': '건강관리장비와서비스', 'Health Care Supplies': '건강관리장비와서비스',
    'Health Care Distributors': '건강관리장비와서비스', 'Health Care Services': '건강관리장비와서비스',
    'Health Care Facilities': '건강관리장비와서비스', 'Managed Health Care': '건강관리장비와서비스',
    'Health Care Technology': '건강관리장비와서비스',
    # 제약과생물공학
    'Pharmaceuticals': '제약과생물공학', 'Biotechnology': '제약과생물공학',
    'Life Sciences Tools & Services': '제약과생물공학',
    # 은행
    'Diversified Banks': '은행', 'Regional Banks': '은행',
    # 증권
    'Investment Banking & Brokerage': '증권',
    # 다각화된금융
    'Multi-Sector Holdings': '다각화된금융', 'Specialized Finance': '다각화된금융',
    'Consumer Finance': '다각화된금융', 'Asset Management & Custody Banks': '다각화된금융',
    'Financial Exchanges & Data': '다각화된금융', 'Transaction & Payment Processing Services': '다각화된금융',
    # 보험
    'Property & Casualty Insurance': '보험', 'Life & Health Insurance': '보험',
    'Multi-line Insurance': '보험', 'Insurance Brokers': '보험', 'Reinsurance': '보험',
    # 부동산
    'Diversified REITs': '부동산', 'Industrial REITs': '부동산', 'Hotel & Resort REITs': '부동산',
    'Office REITs': '부동산', 'Health Care REITs': '부동산', 'Residential REITs': '부동산',
    'Retail REITs': '부동산', 'Specialized REITs': '부동산', 'Telecom Tower REITs': '부동산',
    'Timber REITs': '부동산', 'Data Center REITs': '부동산', 'Self-Storage REITs': '부동산',
    'Real Estate Services': '부동산', 'Real Estate Development': '부동산',
    # 소프트웨어와서비스
    'Systems Software': '소프트웨어와서비스', 'Application Software': '소프트웨어와서비스',
    'IT Consulting & Other Services': '소프트웨어와서비스',
    # 기술하드웨어와장비
    'Technology Hardware, Storage & Peripherals': '기술하드웨어와장비',
    'Communications Equipment': '기술하드웨어와장비', 'Electronic Equipment & Instruments': '기술하드웨어와장비',
    'Electronic Components': '기술하드웨어와장비', 'Electronic Manufacturing Services': '기술하드웨어와장비',
    'Technology Distributors': '기술하드웨어와장비',
    # 반도체와반도체장비
    'Semiconductors': '반도체와반도체장비', 'Semiconductor Materials & Equipment': '반도체와반도체장비',
    # 전자와 전기제품
    'Consumer Electronics': '전자와 전기제품',
    # 전기통신서비스
    'Integrated Telecommunication Services': '전기통신서비스',
    'Alternative Carriers': '전기통신서비스', 'Wireless Telecommunication Services': '전기통신서비스',
    # 미디어와엔터테인먼트
    'Interactive Media & Services': '미디어와엔터테인먼트', 'Movies & Entertainment': '미디어와엔터테인먼트',
    'Interactive Home Entertainment': '미디어와엔터테인먼트',
    'Cable & Satellite': '미디어와엔터테인먼트', 'Publishing': '미디어와엔터테인먼트',
    'Broadcasting': '미디어와엔터테인먼트', 'Advertising': '미디어와엔터테인먼트',
    # 유틸리티
    'Electric Utilities': '유틸리티', 'Gas Utilities': '유틸리티',
    'Multi-Utilities': '유틸리티', 'Water Utilities': '유틸리티',
    'Independent Power Producers & Energy Traders': '유틸리티',
    'Renewable Electricity': '유틸리티',
}

# yfinance industry → WICS
YF_TO_WICS = {
    'Oil & Gas Midstream': '에너지', 'Uranium': '에너지', 'Solar': '에너지',
    'Oil & Gas E&P': '에너지', 'Oil & Gas Integrated': '에너지',
    'Oil & Gas Equipment & Services': '에너지', 'Oil & Gas Refining & Marketing': '에너지',
    'Chemicals': '소재', 'Other Industrial Metals & Mining': '소재',
    'Specialty Chemicals': '소재', 'Building Materials': '소재',
    'Aerospace & Defense': '자본재', 'Conglomerates': '자본재',
    'Engineering & Construction': '자본재', 'Specialty Industrial Machinery': '자본재',
    'Electrical Equipment & Parts': '자본재', 'Building Products & Equipment': '자본재',
    'Farm & Heavy Construction Machinery': '자본재', 'Tools & Accessories': '자본재',
    'Staffing & Employment Services': '상업서비스와공급품',
    'Consulting Services': '상업서비스와공급품', 'Waste Management': '상업서비스와공급품',
    'Security & Protection Services': '상업서비스와공급품',
    'Airlines': '운송', 'Railroads': '운송', 'Trucking': '운송',
    'Integrated Freight & Logistics': '운송', 'Marine Shipping': '운송',
    'Auto Manufacturers': '자동차와부품', 'Auto Parts': '자동차와부품',
    'Apparel Manufacturing': '내구소비재와의류', 'Apparel Retail': '내구소비재와의류',
    'Footwear & Accessories': '내구소비재와의류', 'Luxury Goods': '내구소비재와의류',
    'Leisure': '내구소비재와의류', 'Furnishings, Fixtures & Appliances': '내구소비재와의류',
    'Residential Construction': '내구소비재와의류', 'Home Furnishings & Fixtures': '내구소비재와의류',
    'Restaurants': '호텔,레스토랑,레저등', 'Resorts & Casinos': '호텔,레스토랑,레저등',
    'Lodging': '호텔,레스토랑,레저등', 'Gambling': '호텔,레스토랑,레저등',
    'Discount Stores': '소매(유통)', 'Home Improvement Retail': '소매(유통)',
    'Specialty Retail': '소매(유통)', 'Internet Retail': '소매(유통)',
    'Department Stores': '소매(유통)', 'Auto & Truck Dealerships': '소매(유통)',
    'Grocery Stores': '식품과기본식료품소매', 'Pharmaceutical Retailers': '식품과기본식료품소매',
    'Food Distribution': '식품과기본식료품소매',
    'Beverages - Non-Alcoholic': '식품,음료,담배', 'Tobacco': '식품,음료,담배',
    'Beverages - Brewers': '식품,음료,담배', 'Beverages - Wineries & Distilleries': '식품,음료,담배',
    'Packaged Foods': '식품,음료,담배', 'Confectioners': '식품,음료,담배',
    'Farm Products': '식품,음료,담배',
    'Household & Personal Products': '가정용품과개인용품',
    'Medical Devices': '건강관리장비와서비스', 'Medical Instruments & Supplies': '건강관리장비와서비스',
    'Diagnostics & Research': '건강관리장비와서비스', 'Health Information Services': '건강관리장비와서비스',
    'Medical Care Facilities': '건강관리장비와서비스', 'Medical Distribution': '건강관리장비와서비스',
    'Healthcare Plans': '건강관리장비와서비스',
    'Drug Manufacturers - General': '제약과생물공학',
    'Drug Manufacturers - Specialty & Generic': '제약과생물공학',
    'Biotechnology': '제약과생물공학',
    'Banks - Diversified': '은행', 'Banks - Regional': '은행',
    'Capital Markets': '증권',
    'Asset Management': '다각화된금융', 'Credit Services': '다각화된금융',
    'Financial Data & Stock Exchanges': '다각화된금융',
    'Financial Conglomerates': '다각화된금융',
    'Insurance - Life': '보험', 'Insurance - Property & Casualty': '보험',
    'Insurance - Diversified': '보험', 'Insurance Brokers': '보험',
    'REIT - Diversified': '부동산', 'REIT - Industrial': '부동산',
    'REIT - Residential': '부동산', 'REIT - Retail': '부동산',
    'REIT - Healthcare Facilities': '부동산', 'REIT - Office': '부동산',
    'REIT - Specialty': '부동산', 'Real Estate Services': '부동산',
    'REIT - Hotel & Motel': '부동산',
    'Software - Application': '소프트웨어와서비스',
    'Software - Infrastructure': '소프트웨어와서비스',
    'Information Technology Services': '소프트웨어와서비스',
    'Computer Hardware': '기술하드웨어와장비', 'Communication Equipment': '기술하드웨어와장비',
    'Electronic Components': '기술하드웨어와장비', 'Scientific & Technical Instruments': '기술하드웨어와장비',
    'Semiconductors': '반도체와반도체장비',
    'Semiconductor Equipment & Materials': '반도체와반도체장비',
    'Consumer Electronics': '전자와 전기제품',
    'Telecom Services': '전기통신서비스',
    'Internet Content & Information': '미디어와엔터테인먼트',
    'Entertainment': '미디어와엔터테인먼트',
    'Electronic Gaming & Multimedia': '미디어와엔터테인먼트',
    'Advertising Agencies': '미디어와엔터테인먼트', 'Publishing': '미디어와엔터테인먼트',
    'Broadcasting': '미디어와엔터테인먼트',
    'Utilities - Regulated Electric': '유틸리티',
    'Utilities - Regulated Gas': '유틸리티',
    'Utilities - Regulated Water': '유틸리티',
    'Utilities - Diversified': '유틸리티',
    'Utilities - Independent Power Producers': '유틸리티',
    'Utilities - Renewable': '유틸리티',
}


def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')


def write_row(ws, row_idx, values):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=v)
        cell.border = thin


print("=== 1. 국내 WICS 로드 ===")
with open('wics_mapping.json', encoding='utf-8') as f:
    wics = json.load(f)['mapping']
print(f"  {len(wics)}종목")

print("\n=== 2. S&P 500 로드 (Wikipedia GICS) ===")
headers_req = {'User-Agent': 'Mozilla/5.0'}
r = requests.get('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies', headers=headers_req, timeout=30)
sp500 = pd.read_html(io.StringIO(r.text))[0]
print(f"  {len(sp500)}종목")

sp500_data = []
unmapped_gics = set()
for _, row in sp500.iterrows():
    sym = row['Symbol']
    name = row['Security']
    gics_sector = row['GICS Sector']
    gics_sub = row['GICS Sub-Industry']
    wics_sector = GICS_TO_WICS.get(gics_sub, '')
    if not wics_sector:
        unmapped_gics.add(gics_sub)
    sp500_data.append((sym, name, gics_sector, gics_sub, wics_sector))

if unmapped_gics:
    print(f"  매핑 안 된 GICS Sub-Industry: {unmapped_gics}")
    # yfinance fallback
    for i, (sym, name, gs, gsub, ws) in enumerate(sp500_data):
        if not ws:
            try:
                info = yf.Ticker(sym).info
                ind = info.get('industry', '')
                ws = YF_TO_WICS.get(ind, ind)
                sp500_data[i] = (sym, name, gs, gsub, ws)
            except:
                pass

print("\n=== 3. NASDAQ 시총 상위 500 ===")
r = requests.get('https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=500&offset=0',
                  headers=headers_req, timeout=30)
nasdaq_rows = r.json()['data']['table']['rows']
nasdaq_syms = [r['symbol'] for r in nasdaq_rows if '^' not in r['symbol'] and '/' not in r['symbol']]
sp500_syms = set(sp500['Symbol'])

# S&P 500에 없는 NASDAQ 종목만
nasdaq_only = [s for s in nasdaq_syms if s not in sp500_syms]
print(f"  전체 {len(nasdaq_syms)}, S&P500 제외 {len(nasdaq_only)}")

nasdaq_data = []
for i, sym in enumerate(nasdaq_only):
    try:
        info = yf.Ticker(sym).info
        name = info.get('shortName', info.get('longName', ''))
        industry = info.get('industry', '')
        wics_sector = YF_TO_WICS.get(industry, industry)
        nasdaq_data.append((sym, name, industry, wics_sector))
    except:
        nasdaq_data.append((sym, '', '', ''))
    if (i + 1) % 50 == 0:
        print(f"  {i+1}/{len(nasdaq_only)} 처리...")

print(f"  {len(nasdaq_data)}종목 수집 완료")

print("\n=== 4. 엑셀 생성 ===")
wb = openpyxl.load_workbook('WICS_분류.xlsx')

# --- 최종본 시트 ---
if '최종본' in wb.sheetnames:
    del wb['최종본']
ws = wb.create_sheet('최종본')
write_header(ws, ['구분', '티커/코드', '기업명', 'WICS 소분류', '원문 분류'])

row_idx = 2

# 국내
for code, wics_name in sorted(wics.items(), key=lambda x: x[1]):
    write_row(ws, row_idx, ['KRW', code, '', wics_name, ''])
    row_idx += 1

# S&P 500
for sym, name, gs, gsub, ws_name in sp500_data:
    write_row(ws, row_idx, ['S&P500', sym, name, ws_name, gsub])
    row_idx += 1

# NASDAQ only
for sym, name, ind, ws_name in nasdaq_data:
    write_row(ws, row_idx, ['NASDAQ', sym, name, ws_name, ind])
    row_idx += 1

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 40
ws.auto_filter.ref = f'A1:E{row_idx-1}'

# --- 분류 매핑 시트 ---
if '분류 매핑' in wb.sheetnames:
    del wb['분류 매핑']
ws_map = wb.create_sheet('분류 매핑', 0)
mapping_display = [
    ('에너지', 'Oil & Gas, Uranium, Solar, Coal'),
    ('소재', 'Chemicals, Metals & Mining, Steel, Construction Materials, Paper'),
    ('자본재', 'Aerospace & Defense, Conglomerates, Construction, Machinery, Electrical Equipment'),
    ('상업서비스와공급품', 'Environmental Services, HR, Consulting, Data Processing'),
    ('운송', 'Airlines, Railroads, Trucking, Marine, Logistics'),
    ('자동차와부품', 'Auto Manufacturers, Auto Parts, Tires'),
    ('내구소비재와의류', 'Apparel, Footwear, Luxury, Leisure, Home Furnishings, Homebuilding'),
    ('호텔,레스토랑,레저등', 'Hotels, Restaurants, Casinos, Leisure Facilities'),
    ('소매(유통)', 'Internet Retail, Discount/Specialty/Home Improvement Retail'),
    ('교육서비스', '—'),
    ('식품과기본식료품소매', 'Food/Drug Retail, Food Distribution'),
    ('식품,음료,담배', 'Packaged Foods, Beverages, Tobacco, Farm Products'),
    ('가정용품과개인용품', 'Household & Personal Products'),
    ('건강관리장비와서비스', 'Medical Devices/Instruments, Health Care Equipment/Services/Facilities'),
    ('제약과생물공학', 'Pharmaceuticals, Biotechnology, Life Sciences'),
    ('은행', 'Diversified/Regional Banks'),
    ('증권', 'Investment Banking, Capital Markets'),
    ('다각화된금융', 'Asset Management, Consumer Finance, Financial Exchanges, Payment Processing'),
    ('보험', 'Insurance (Life, P&C, Diversified, Brokers, Reinsurance)'),
    ('부동산', 'REITs (all types), Real Estate Services'),
    ('소프트웨어와서비스', 'Software (Application/Infrastructure), IT Services/Consulting'),
    ('기술하드웨어와장비', 'Computer Hardware, Communication Equipment, Electronic Components/Instruments'),
    ('반도체와반도체장비', 'Semiconductors, Semiconductor Equipment & Materials'),
    ('전자와 전기제품', 'Consumer Electronics'),
    ('디스플레이', '—'),
    ('전기통신서비스', 'Telecom Services (Integrated, Wireless, Alternative)'),
    ('미디어와엔터테인먼트', 'Interactive Media, Movies, Gaming, Broadcasting, Publishing, Advertising'),
    ('유틸리티', 'Electric/Gas/Water Utilities, Independent Power, Renewable'),
]
write_header(ws_map, ['WICS 소분류', 'GICS / yfinance 대응 항목'])
for i, (w, g) in enumerate(mapping_display, 2):
    ws_map.cell(row=i, column=1, value=w).border = thin
    ws_map.cell(row=i, column=1).font = Font(bold=True)
    ws_map.cell(row=i, column=2, value=g).border = thin

ws_map.column_dimensions['A'].width = 24
ws_map.column_dimensions['B'].width = 80

wb.save('WICS_분류.xlsx')
total = len(wics) + len(sp500_data) + len(nasdaq_data)
print(f"\n완료! 최종본 {total}종목 (KRW {len(wics)} + S&P500 {len(sp500_data)} + NASDAQ {len(nasdaq_data)})")
print(f"시트: {wb.sheetnames}")
